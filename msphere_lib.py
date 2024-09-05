# -*- coding: utf-8 -*-
'''
@author: olivia.dou
Created on: 2023/6/28 15:44
desc:
'''
import os,re,random,string
from openpyxl import load_workbook

def convert_path(path):
    converted_path = re.sub(r'\\', r'\\\\', path)
    return converted_path

def get_parts_from_comment_field(field_text, image_dir):
    """从评论字段值中获取文本、图片部分
    导出的评论字段可能包含多条评论记录，每条评论记录可能包含文本部分和图片部分（共三部分，图片前文本、图片路径、图片后文本）

    :param field_text: 字段的完整文本
    :param image_dir: 图片的目录路径

    :return:
    """

    results = []
    try:
        field_text = re.sub('【.+】\n', '', field_text)
        if field_text is not None and field_text != '' and "![image.png](/resource/md/get?fileName=" in field_text:
            field_text = field_text.replace("![image.png](/resource/md/get?fileName=", image_dir + os.sep)
            field_text = field_text.replace(")", "")

        #result_str = actual_result
        for row_str in field_text.split('\n'):
            image_path_pattern = convert_path(image_dir + os.sep) + r'.+\.png'


            regex = re.compile(image_path_pattern)
            match = regex.search(row_str)

            if match: #有图片
                start_index = match.start()
                end_index = match.end()

                prefix = row_str[:start_index]
                image_path = row_str[start_index:end_index]
                suffix = row_str[end_index:]

                results.append((prefix,image_path,suffix))
            else:
                results.append((row_str,'',''))

        return results
    except Exception as e:
        return []

def generate_random_string(length):
    letters = string.ascii_letters + string.digits  # 包含字母和数字
    random_string = ''.join(random.choice(letters) for _ in range(length))
    return random_string


def get_merged_range(sheet, cell):

    # 遍历合并单元格范围
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range

    return None


def get_tc_data(tc_file_path, src_sheet_name):
    """读取metersphere导出格式的用例数据

    :param tc_file:
    :return:
    """
    src_wb = load_workbook(tc_file_path)
    src_sheet = src_wb[src_sheet_name]
    head_row = src_sheet[1]
    columns = [cell.value for cell in head_row if cell.value is not None]
    tc_data = []
    tc_row = {}
    num_rows = 0
    previous_module_paths = []

    try:
        for i in range(1, src_sheet.max_row):
            # 读取模块信息
            # module_cell = src_sheet.cell(row=i+1,column=3)
            module_cell = src_sheet.cell(row=i + 1, column=columns.index('所属模块') + 1)

            mcr = get_merged_range(src_sheet, module_cell)
            if mcr is not None and num_rows == 0:  # 合并单元格
                case_rows = mcr.max_row - mcr.min_row + 1
                num_rows = case_rows
            elif mcr is None:
                num_rows = 1
            module = module_cell.value
            # 读取步骤

            if module is not None and module != '':
                module_paths = module.strip('/').split('/')
                # module_levels = [(i,module) for i,module in enumerate(module_paths)]
                # actual_result = src_sheet.cell(row=i + 1, column=columns.index('附加字段')+1).value
                actual_result = src_sheet.cell(row=i + 1, column=columns.index('评论') + 1).value
                # actual_result = actual_result.strip('\n')
                # actual_result = re.sub('【.+】', '', actual_result)
                # if actual_result is not None and actual_result!='' and "![image.png](/resource/md/get?fileName=" in actual_result:
                #     actual_result = actual_result.replace("![image.png](/resource/md/get?fileName=", image_dir + os.sep)
                #     actual_result = actual_result.replace(")", "")
                # print(actual_result)

                test_step = src_sheet.cell(row=i + 1, column=columns.index('步骤描述') + 1).value
                test_step = "" if test_step is None else test_step
                expected_result = src_sheet.cell(row=i + 1, column=columns.index('预期结果') + 1).value
                expected_result = "" if expected_result is None else expected_result
                tc_row = {"test_case_id": src_sheet.cell(row=i + 1, column=columns.index('ID') + 1).value,
                          "test_case_name": src_sheet.cell(row=i + 1, column=columns.index('用例名称') + 1).value,
                          "precondition": src_sheet.cell(row=i + 1, column=columns.index('前置条件') + 1).value,
                          "test_purpose": src_sheet.cell(row=i + 1, column=columns.index('备注') + 1).value,
                          "test_steps": "步骤1. " + test_step,
                          "expected_result": "步骤1. " + expected_result,
                          "label": src_sheet.cell(row=i + 1, column=columns.index('标签') + 1).value,
                          "test_designer": src_sheet.cell(row=i + 1, column=columns.index('创建人') + 1).value,
                          "test_executor": src_sheet.cell(row=i + 1, column=columns.index('责任人') + 1).value,
                          "tc_level": src_sheet.cell(row=i + 1, column=columns.index('用例等级') + 1).value,
                          "tc_status": src_sheet.cell(row=i + 1, column=columns.index('用例状态') + 1).value,
                          "tc_type": src_sheet.cell(row=i + 1, column=columns.index('用例类型') + 1).value,
                          "actual_result": actual_result,
                          "test_date": str(src_sheet.cell(row=i + 1, column=columns.index('更新时间') + 1).value)[:10]
                          }
                for level, module in enumerate(module_paths):
                    if len(previous_module_paths) >= len(module_paths) and module_paths[level] == previous_module_paths[
                        level]:
                        tc_row['level_%s_module' % level] = module = None
                    else:
                        tc_row['level_%s_module' % level] = module
                num_rows -= 1

            else:  # 合并单元格合并行值为空
                tc_row['test_steps'] = tc_row['test_steps'] + '\n' + "步骤%s. " % (
                            case_rows + 1 - num_rows) + src_sheet.cell(row=i + 1,
                                                                       column=columns.index('步骤描述') + 1).value
                expected_result = src_sheet.cell(row=i + 1, column=columns.index('预期结果') + 1).value
                if expected_result != '' and expected_result is not None:
                    tc_row['expected_result'] = tc_row['expected_result'] + '\n' + "步骤%s. " % (
                                case_rows + 1 - num_rows) + expected_result

                num_rows -= 1

            previous_module_paths = module_paths
            #        print(os.path.isfile(tc_row['actual_result']))
            if num_rows == 0:
                tc_data.append(tc_row)

        return tc_data
    except Exception as e:
        print(e)
        return None
