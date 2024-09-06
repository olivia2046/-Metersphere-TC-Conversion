# -*- coding: utf-8 -*-
'''
@author: olivia.dou
Created on: 2024/1/3 16:50
desc: 将metersphere标准格式的用例，按照指定模板的格式填入sheet中（一个用例一个sheet）
使用方法：
如果导出用例需要附带截图，则本脚本需在Metersphere所在服务器执行（否则找不到图片）
如果不需要附带截图，则脚本可在本地执行
1. Metersphere中按顺序（如用例ID）排列全部用例，全部导出，勾选全部基础字段、自定义字段，以及其他字段中的评论、创建人、创建时间（排序用）、更新时间等
2. 检查导出的Excel用例的顺序，如果和预期输出的用例顺序不一致，手动调整（也可在生成Excel后调整Excel文档中用例顺序）
3. 修改rootdir （脚本中使用当前文件路径，可修改为客户项目文件夹)，rootdir下创建好tmpl_wb_path指向的用例模板文件，参加Excel分sheet用例模板.xlsx
4. 设置src_wb_path
5. 设置tmpl_sheet_name和tmpl_sheet_name
6. 修改actual_result_loc（实际结果内容在表格中的位置索引，如A7）
7. 设置need_image(若输出的用例不需要带截图，则设置为False)。需要输出截图时注意设置img_width和img_height
8. 修改image_dir(用例截图所在路径，服务器上为固定地址，仅调试脚本时需要修改)
9. 使用python3运行脚本

'''
import os
from shutil import copyfile
from openpyxl.drawing.image import Image
from openpyxl import load_workbook, Workbook
from msphere_lib import get_tc_data,get_parts_from_comment_field


# field_mapping = {"ID":"test_case_id","用例名称":"test_case_name","步骤描述":"test_steps","预期结果":"expected_result",
#                  "创建人":"test_designer","责任人":"test_executor","更新时间":"test_date","评论":""}
current_file_path = os.path.abspath(__file__)
root_dir = os.path.dirname(current_file_path)

# 是否导出图片
need_image = False
image_dir = r"/app/metersphere/data/image/markdown"
img_width = 600
img_height = 450

src_wb_path = root_dir + os.sep + "Metersphere_case_dummy.xlsx"
src_sheet_name = "模版" #指定源用例文件中的sheet名
tmpl_wb_path = root_dir + os.sep + "Excel分sheet用例模板.xlsx"
tmpl_sheet_name = "测试用例执行结果" # 指定模版sheet名
sheet_name_prefix="测试用例执行结果" # 指定生成的测试用例sheet名前缀
#实际结果在模板中的位置
actual_result_loc = 'A7'
target_wb_path = root_dir + os.sep + "excel_multi_sheets_target.xlsx"
# 因需要带格式复制sheet， 且copy_worksheet只能在同一个workbook内复制sheet，因此先复制模板文件到目标文件，再在目标文件内操作
try:
    copyfile(tmpl_wb_path, target_wb_path)
    #target_wb = Workbook()
    target_wb = load_workbook(target_wb_path)
except Exception as e:
    print(e)


# Todo: 判断target_wb是否已打开，若是则提示并退出

def replace_string_in_excel(input_file, sheet_name, case_data):
    """

    :param input_file:
    :param output_file:
    :param case_data:
    :param field_mapping:
    :return:
    """
    # # 加载Excel文件
    # workbook = load_workbook(input_file)
    # sheet = workbook[tmpl_sheet_name]
    wb = load_workbook(input_file)
    tmpl_sheet = wb[sheet_name]
    #target_wb = load_workbook(output_file)

    new_sheet = wb.copy_worksheet(tmpl_sheet)
    new_sheet.title = sheet_name_prefix + case_data['test_case_id']
    
    for k,v in case_data.items():
        # 遍历每个单元格
        for row in new_sheet.iter_rows():
            for cell in row:
                if cell.value is not None and "«%s»"%k in str(cell.value):
                    # 替换字符串
                    cell.value = str(cell.value).replace("«%s»" % k, case_data[k])

    actual_results = get_parts_from_comment_field(case_data['actual_result'],image_dir)
    #插入实际结果
    for result in actual_results:
        image_path = result[1]

        if len(image_path) > 0 and need_image:
            img = Image(image_path)

            # 设置图片的位置（A1单元格的左上角作为起点）
            new_sheet.add_image(img, actual_result_loc)
            img.width = img_width
            img.height = img_height

        #run.add_text(result[2])


    wb.save(input_file)

tmpl_wb = load_workbook(tmpl_wb_path)
tmpl_sheet = tmpl_wb[tmpl_sheet_name]

tc_data = get_tc_data(src_wb_path, src_sheet_name)

for row in tc_data:
    # 调用函数，指定输入和输出文件路径，以及替换的字符串
    replace_string_in_excel(target_wb_path, tmpl_sheet_name, row)
    # 保存修改后的Excel文件
    #target_wb = Workbook()
    #target_wb[sheet_name] =

#workbook.save(target_wb)
